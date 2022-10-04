import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList

import string
import os


input_file = 'input_data/supermarket_sales.xlsx'
output_file = 'output_data/report_2019.xlsx'

# Read Excel file
df = pd.read_excel(input_file)
df[['Gender', 'Product line', 'Total']].tail()

# Make pivot table
df1 = df.pivot_table(index='Gender',
                    columns='Product line',
                    values='Total',
                    aggfunc='sum').round()
print(df1)

# send the report table to excel file
df1.to_excel(output_file,
            sheet_name='Product line',
            startrow=4)

wb = load_workbook(output_file)
wb.active = wb['Product line']
sheet = wb['Product line']
# cell reference
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print(min_column, max_column, min_row, max_row)

# Chart Excel
barchart = BarChart()

data = Reference(sheet,
                min_col=min_column+1,
                max_col=max_column,
                min_row=min_row,
                max_row=max_row)
categories = Reference(sheet,
                    min_col=min_column,
                    max_col=max_column,
                    min_row=min_row+1,
                    max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
   
sheet.add_chart(barchart, 'B12')
barchart.title = 'Sales by Product line'
barchart.style = 2
wb.save(output_file)

# Menambah column Total
alphabet = list(string.ascii_uppercase)
alphabet_excel = alphabet[:max_column]

for i in alphabet_excel:
    if i != 'A':
        sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        sheet[f'{i}{max_row+1}'].style = 'Currency'
sheet[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

sheet['A1'] = 'Sales Report'
sheet['A2'] = '2019'
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=12)
wb.save(output_file)


# Workbook 2
df2 = df[['City', 'Total']].groupby(['City']).sum().round()

df2['Percent'] = ((df2['Total'] / df2['Total'].sum()) * 100).round()

del df2['Total']
print(df2)

df2.to_excel(output_file,
            sheet_name='City',
            startrow=4)

with pd.ExcelWriter(output_file) as writer:
    df1.to_excel(writer, sheet_name='Product line', startrow=4)
    df2.to_excel(writer, sheet_name='City', startrow=4)
    
wb = load_workbook(output_file)
wb.active = wb['City']
sheet_2 = wb['City']

min_column_2 = sheet_2.min_column
max_column_2 = sheet_2.max_column
min_row_2 = sheet_2.min_row
max_row_2 = sheet_2.max_row

print(min_column_2, max_column_2, min_row_2, max_row_2)

piechart = PieChart3D()

data_2 = Reference(sheet_2,
                    min_col=min_column_2+1,
                    max_col=max_column_2,
                    min_row=min_row_2,
                    max_row=max_row_2)
categories_2 = Reference(sheet_2,
                        min_col=min_column_2,
                        max_col=max_column_2,
                        min_row=min_row_2+1,
                        max_row=max_row_2)

piechart.add_data(data_2, titles_from_data=True)
piechart.set_categories(categories_2)
piechart.dataLabels = DataLabelList()
piechart.dataLabels.showVal = True

sheet_2.add_chart(piechart, 'B12')
piechart.title = 'Sales by Region'
piechart.style = 2

alphabet = list(string.ascii_uppercase)
alphabet_excel = alphabet[:max_column]

sheet_2['A1'] = 'Sales Report'
sheet_2['A2'] = '2019'
sheet_2['A1'].font = Font('Arial', bold=True, size=20)
sheet_2['A2'].font = Font('Arial', bold=True, size=12)
wb.save(output_file)

wb.save(output_file)


def automate_excel(file_name, sheet_name, index, chart_style, columns=None):

    excel_file = pd.read_excel(input_file)
    df1 = excel_file.pivot_table(index='Gender',
                                columns='Product line',
                                values='Total',
                                aggfunc='sum').round()

    try:
        with pd.ExcelWriter(output_file, mode='a', if_sheet_exists='replace') as writer:
            df1.to_excel(writer, sheet_name='Product line', startrow=4)
    except KeyError:
        with pd.ExcelWriter(output_file) as writer:
            df1.to_excel(writer, sheet_name='Product line', startrwo=4)

    wb = load_workbook(output_file)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.active = wb['Product line']
    sheet = wb['Product line']

    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    barchart = BarChart()

    data = Reference(wb.active,
                    min_col=min_column+1,
                    max_col=max_column,
                    min_row=min_row,
                    max_row=max_row)
    categories = Reference(wb.active,
                        min_col=min_column,
                        max_col=max_column,
                        min_row=min_row+1,
                        max_row=max_row)

    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)

    sheet.add_chart(barchart, 'B12')
    barchart.title = 'Sales by Product line'
    barchart.style = 2
    wb.save(output_file)

    alphabet = list(string.ascii_uppercase)
    alphabet_excel = alphabet[:max_column]

    for i in alphabet_excel:
        if i != 'A':
            sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
            sheet[f'{i}{max_row+1}'].style = 'Currency'
    sheet[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

    sheet['A1'] = 'Sales Report'
    sheet['A2'] = '2019'
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=12)
    wb.save(output_file)

automate_excel(output_file, 'Product line', 'Gender', BarChart(), 'Product line')
automate_excel(output_file, 'City', 'City', PieChart3D())