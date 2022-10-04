# membuat Automated Report (table dan grafik), dan dikirim ke discord

import pandas as pd #pandas untuk membuat dataframe(df)
from openpyxl import workbook #untuk berinteraksi antara python dan file excel
import openpyxl 
from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList

import string 
import os 

input_file = 'input_data/supermarket_sales.xlsx'
output_file = 'output_data/report_penjualan_2019.xlsx'

# PART I - Load DataSet
df = pd.read_excel(input_file)

# print kolom dan sample 5 pertama 


# penjualan total per gender dan product line
print(df[['Gender', 'Product line', 'Total']].head())

# memilih salah satu gender
# print(df.loc[df['Gender'] == 'Male'].head())
# print(df['Gender'].loc[df['Gender'] == 'Male'].head())

# membuat pivot (merubah salah satu row menjadi kolom) table

df = df.pivot_table(index = 'Gender', 
                    columns= 'Product line', 
                    values= 'Total', 
                    aggfunc='sum'
                    ).round()

# print(f' DataFrame columns: {df.columns}')
# print(f'Sample Dataset: {df.head().to_string()}')

print('Save DataFrame to Excel . . .')

df.to_excel(output_file, 
                sheet_name='Report', 
                startrow= 4
            )

print('Save to Excel done!')

# PART II - Grafik
wb = load_workbook(output_file)
wb.active = wb['Report']
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print('''
min_column: {0}
max_column: {1}
min_row: {2}
max_row: {3}
    '''.format(min_column, max_column, min_row, max_row))

barchart = BarChart()

data = Reference(sheet,
                min_col = min_column+1,
                max_col = max_column,
                min_row = min_row,
                max_row = max_row
                )

categories = Reference(sheet,
                        min_col = min_column,
                        max_col = max_column,
                        min_row = min_row+1,
                        max_row = max_row
                        )

barchart.add_data(data, titles_from_data = True)
barchart.set_categories(categories)

sheet.add_chart(barchart, 'B12')
barchart.title = 'Sales Berdasarkan Product'
barchart.style = 2
# wb.save(output_file)


# Total dari penjualan
alphabet = list(string.ascii_uppercase)
alphabet_excel = alphabet[:max_column]
# [A,B,C,D,E,F,G]
for i in alphabet_excel:
    if i != 'A':
        wb.active[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        wb.active[f'{i}{max_row+1}'].style = 'Currency'

wb.active[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

sheet['A1'] = 'Sales Report'
sheet['A2'] = '2019'
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=12)

wb.save(output_file)

# send to discord
import requests
import discord
from discord import SyncWebhook

webhook_url = 'https://discordapp.com/api/webhooks/1025306462551031829/OP6jPfNuO6A92p7QJvPzBvOViFVZSZ7TivnFJygzaD5DZLx4Yie4Bng2FVZoQ4yj0Qzo'

webhook = SyncWebhook.from_url(webhook_url)

with open(file=output_file, mode='rb') as file:
    excel_file = discord.File(file)

webhook.send('This is an automated message',
            username='Bot',
            file=excel_file
            )
